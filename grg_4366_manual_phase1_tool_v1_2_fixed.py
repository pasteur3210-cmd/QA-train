
import logging
import queue
import re
import shutil
import socket
import telnetlib
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Tuple, List

import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


APP_VERSION = "Phase1-V1.2"
APP_TITLE = "GRG-4366 第一階人工查核工具"
HOST = "192.168.1.1"
PORT = 23
TIMEOUT = 12
ENCODING = "ascii"
USERNAME = "root"
PASSWORD = "12345"
LABELINFO_CMD = "labelinfo"

INITIAL_READ_DELAY = 1.0
READ_INTERVAL = 0.15
READ_IDLE_BREAK = 4

RESULT_OPTIONS = ["", "PASS", "FAIL", "N/A"]

BASE_DIR = Path.cwd()
LOG_DIR = BASE_DIR / "logs"
OUTPUT_DIR = BASE_DIR / "output"
BACKUP_DIR = BASE_DIR / "backup"


class TelnetSessionError(Exception):
    pass


class CommandTimeoutError(TelnetSessionError):
    pass


def ensure_dirs() -> None:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)


def get_daily_paths() -> Tuple[Path, Path]:
    date_tag = datetime.now().strftime("%Y%m%d")
    excel_path = OUTPUT_DIR / f"grg_4366_phase1_inspection_{date_tag}.xlsx"
    log_path = LOG_DIR / f"grg_4366_phase1_inspection_{date_tag}.log"
    return excel_path, log_path


ensure_dirs()
EXCEL_FILE, LOG_FILE = get_daily_paths()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_FILE, encoding="utf-8"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)


def sanitize_excel_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()


def normalize_mac(value: str) -> str:
    v = re.sub(r"[^0-9A-Fa-f]", "", value or "")
    if len(v) != 12:
        return (value or "").strip().upper()
    return ":".join(v[i:i+2] for i in range(0, 12, 2)).upper()


def normalize_sn(value: str) -> str:
    return (value or "").strip().upper()


def normalize_text(value: str) -> str:
    return (value or "").strip()


def normalize_gpon_sn(value: str) -> str:
    return (value or "").strip().upper()


def validate_sn(value: str) -> bool:
    value = (value or "").strip()
    return bool(value and re.fullmatch(r"[A-Za-z0-9\-_]+", value))


def validate_mac(value: str) -> bool:
    value = (value or "").strip()
    return bool(re.fullmatch(r"([0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}|[0-9A-Fa-f]{12}", value))


def validate_gpon_sn(value: str) -> bool:
    value = (value or "").strip()
    return bool(re.fullmatch(r"[A-Za-z0-9\-_]+", value))


def parse_qrcode(raw: str) -> Dict[str, str]:
    text = sanitize_excel_text(raw)
    if not text:
        return {"qr_sn": "", "qr_mac": "", "qr_wpa": ""}

    compact_text = re.sub(r"\s+", " ", text).strip()
    parsed = {"qr_sn": "", "qr_mac": "", "qr_wpa": ""}

    sn_patterns = [
        r"(?i)\bS/?N\s*[:=]\s*([A-Za-z0-9\-_]+?)(?=\s*(?:MAC|WPA|WIFI)\s*[:=]|$)",
        r"(?i)\bSN\s*[:=]\s*([A-Za-z0-9\-_]+?)(?=\s*(?:MAC|WPA|WIFI)\s*[:=]|$)",
    ]
    mac_patterns = [
        r"(?i)\bMAC\s*[:=]\s*([0-9A-Fa-f]{12})(?=\s*(?:WPA|WIFI|S/?N|SN)\s*[:=]|$)",
        r"(?i)\bMAC\s*[:=]\s*([0-9A-Fa-f:\-]{12,17})(?=\s*(?:WPA|WIFI|S/?N|SN)\s*[:=]|$)",
    ]
    wpa_patterns = [
        r"(?i)\bWPA(?:\s*KEY)?\s*[:=]\s*([^\s]+)",
        r"(?i)\bWIFI(?:\s*KEY)?\s*[:=]\s*([^\s]+)",
    ]

    for p in sn_patterns:
        m = re.search(p, compact_text)
        if m:
            parsed["qr_sn"] = normalize_sn(m.group(1))
            break

    for p in mac_patterns:
        m = re.search(p, compact_text)
        if m:
            parsed["qr_mac"] = normalize_mac(m.group(1))
            break

    for p in wpa_patterns:
        m = re.search(p, compact_text)
        if m:
            parsed["qr_wpa"] = m.group(1).strip()
            break

    return parsed

LABELINFO_KEY_MAP = {
    "VERSION": "version",
    "MAC": "mac_dut",
    "SN": "sn_dut",
    "5G SSID": "ssid_5g",
    "5G WPAKEY": "wpa_5g",
    "2.4G SSID": "ssid_24g",
    "2.4G WPAKEY": "wpa_24g",
    "MLO SSID": "mlo_ssid",
    "MLO WPAKEY": "mlo_wpa",
    "GPON_SN": "gpon_sn_dut",
    "GPON PLOAM PASSWD": "gpon_ploam_passwd",
    "LOID": "loid",
    "LOID PASSWD": "loid_passwd",
    "PON VENDOR ID": "pon_vendor_id",
    "HW_HWWER": "hw_hwwer",
    "HW CWMP PRODUCTCLASS": "hw_cwmp_productclass",
    "GPON ONU MODEL": "gpon_onu_model",
    "PON BID": "pon_bid",
}


def normalize_labelinfo_key(key: str) -> str:
    key = sanitize_excel_text(key).upper().replace("_", " ").strip()
    key = re.sub(r"\s+", " ", key)
    return key


def parse_labelinfo(raw_text: str) -> Dict[str, str]:
    result = {
        "version": "",
        "sn_dut": "",
        "mac_dut": "",
        "gpon_sn_dut": "",
        "ssid_5g": "",
        "wpa_5g": "",
        "ssid_24g": "",
        "wpa_24g": "",
        "mlo_ssid": "",
        "mlo_wpa": "",
        "gpon_ploam_passwd": "",
        "loid": "",
        "loid_passwd": "",
        "pon_vendor_id": "",
        "hw_hwwer": "",
        "hw_cwmp_productclass": "",
        "gpon_onu_model": "",
        "pon_bid": "",
        "raw_text": sanitize_excel_text(raw_text),
    }
    for line in sanitize_excel_text(raw_text).split("\n"):
        if "=" not in line:
            continue
        raw_key, raw_value = line.split("=", 1)
        nkey = normalize_labelinfo_key(raw_key)
        target = LABELINFO_KEY_MAP.get(nkey)
        if target:
            result[target] = raw_value.strip()
    if result["mac_dut"]:
        result["mac_dut"] = normalize_mac(result["mac_dut"])
    if result["sn_dut"]:
        result["sn_dut"] = normalize_sn(result["sn_dut"])
    if result["gpon_sn_dut"]:
        result["gpon_sn_dut"] = normalize_gpon_sn(result["gpon_sn_dut"])
    return result


def compare_fields(label_data: Dict[str, str], dut_data: Dict[str, str], expected_version: str = "") -> Dict[str, str]:
    label_sn = normalize_sn(label_data.get("sn_label", ""))
    label_mac = normalize_mac(label_data.get("mac_label", ""))
    label_gpon = normalize_gpon_sn(label_data.get("gpon_sn_label", ""))

    dut_sn = normalize_sn(dut_data.get("sn_dut", ""))
    dut_mac = normalize_mac(dut_data.get("mac_dut", ""))
    dut_gpon = normalize_gpon_sn(dut_data.get("gpon_sn_dut", ""))
    dut_version = normalize_text(dut_data.get("version", ""))
    expected_version = normalize_text(expected_version)

    sn_match = "PASS" if label_sn and dut_sn and label_sn == dut_sn else "FAIL"
    mac_match = "PASS" if label_mac and dut_mac and label_mac == dut_mac else "FAIL"
    gpon_match = "PASS" if label_gpon and dut_gpon and label_gpon == dut_gpon else "FAIL"

    if expected_version:
        version_auto_check = "PASS" if expected_version == dut_version else "FAIL"
    else:
        version_auto_check = "N/A"

    auto_fail_reasons = []
    if sn_match == "FAIL":
        auto_fail_reasons.append("SN 不一致")
    if mac_match == "FAIL":
        auto_fail_reasons.append("MAC 不一致")
    if gpon_match == "FAIL":
        auto_fail_reasons.append("GPON SN 不一致")
    if version_auto_check == "FAIL":
        auto_fail_reasons.append("Version 不一致")

    auto_result = "PASS" if not auto_fail_reasons else "FAIL"
    return {
        "sn_match": sn_match,
        "mac_match": mac_match,
        "gpon_sn_match": gpon_match,
        "version_auto_check": version_auto_check,
        "auto_result": auto_result,
        "auto_fail_reason": "；".join(auto_fail_reasons),
    }


def build_record(
    *,
    work_order: str,
    inspector: str,
    model_name: str,
    station_id: str,
    remark: str,
    label_data: Dict[str, str],
    dut_data: Dict[str, str],
    compare_data: Dict[str, str],
    manual_data: Dict[str, str],
    final_result: str,
    fail_reason: str,
    session_id: str,
    time_text: Optional[str] = None,
) -> Dict[str, str]:
    manual_summary = manual_data.get("manual_result", "")
    return {
        "Time": time_text or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Session_ID": session_id,
        "Work_Order": work_order,
        "Inspector": inspector,
        "Model": model_name,
        "Station": station_id,
        "Remark": remark,
        "SN_Label": label_data.get("sn_label", ""),
        "MAC_Label": label_data.get("mac_label", ""),
        "GPON_SN_Label": label_data.get("gpon_sn_label", ""),
        "QR_Raw": label_data.get("qr_raw", ""),
        "QR_SN": label_data.get("qr_sn", ""),
        "QR_MAC": label_data.get("qr_mac", ""),
        "QR_WPA": label_data.get("qr_wpa", ""),
        "Version_DUT": dut_data.get("version", ""),
        "SN_DUT": dut_data.get("sn_dut", ""),
        "MAC_DUT": dut_data.get("mac_dut", ""),
        "GPON_SN_DUT": dut_data.get("gpon_sn_dut", ""),
        "SSID_5G": dut_data.get("ssid_5g", ""),
        "WPA_5G": dut_data.get("wpa_5g", ""),
        "SSID_24G": dut_data.get("ssid_24g", ""),
        "WPA_24G": dut_data.get("wpa_24g", ""),
        "MLO_SSID": dut_data.get("mlo_ssid", ""),
        "MLO_WPA": dut_data.get("mlo_wpa", ""),
        "SN_Match": compare_data.get("sn_match", ""),
        "MAC_Match": compare_data.get("mac_match", ""),
        "GPON_SN_Match": compare_data.get("gpon_sn_match", ""),
        "Version_Auto_Check": compare_data.get("version_auto_check", ""),
        "Auto_Result": compare_data.get("auto_result", ""),
        "Auto_Fail_Reason": compare_data.get("auto_fail_reason", ""),
        "Version_Manual": manual_data.get("version_manual", ""),
        "SSID_5G_Manual": manual_data.get("ssid_5g_manual", ""),
        "WPA_5G_Manual": manual_data.get("wpa_5g_manual", ""),
        "SSID_24G_Manual": manual_data.get("ssid_24g_manual", ""),
        "WPA_24G_Manual": manual_data.get("wpa_24g_manual", ""),
        "MLO_SSID_Manual": manual_data.get("mlo_ssid_manual", ""),
        "MLO_WPA_Manual": manual_data.get("mlo_wpa_manual", ""),
        "Manual_Result": manual_summary,
        "Manual_Fail_Reason": manual_data.get("manual_fail_reason", ""),
        "Final_Result": final_result,
        "Fail_Reason": fail_reason,
        "Raw_LabelInfo": dut_data.get("raw_text", ""),
    }


class GRG4366TelnetClient:
    def __init__(self, host: str = HOST, port: int = PORT, timeout: int = TIMEOUT):
        self.host = host
        self.port = port
        self.timeout = timeout
        self.tn: Optional[telnetlib.Telnet] = None

    def connect(self) -> None:
        try:
            logger.info("Connecting to %s:%s", self.host, self.port)
            self.tn = telnetlib.Telnet(self.host, self.port, self.timeout)
        except Exception as exc:
            raise TelnetSessionError(f"無法連線到 {self.host}:{self.port}，原因：{exc}") from exc

    def close(self) -> None:
        if self.tn:
            try:
                self.tn.close()
            except Exception:
                pass
            self.tn = None

    def _read_banner(self) -> str:
        if not self.tn:
            raise TelnetSessionError("Telnet 尚未連線")
        time.sleep(INITIAL_READ_DELAY)
        chunks: List[bytes] = []
        idle = 0
        while idle < READ_IDLE_BREAK:
            try:
                part = self.tn.read_very_eager()
            except EOFError:
                break
            except OSError as exc:
                raise TelnetSessionError(f"讀取初始畫面失敗：{exc}") from exc
            if part:
                chunks.append(part)
                idle = 0
            else:
                idle += 1
            time.sleep(READ_INTERVAL)
        return b"".join(chunks).decode(ENCODING, errors="ignore")

    def _send_line_and_wait(self, cmd: str, timeout: float, settle: float = 0.0, expect_tokens: Optional[List[bytes]] = None) -> str:
        if not self.tn:
            raise TelnetSessionError("Telnet 尚未連線")
        expect_tokens = expect_tokens or []
        try:
            self.tn.write((cmd + "\r\n").encode(ENCODING))
        except (EOFError, BrokenPipeError, ConnectionResetError, ConnectionAbortedError, socket.error, OSError) as exc:
            raise TelnetSessionError(f"送出指令時連線中斷：{cmd} | 原因：{exc}") from exc

        deadline = time.time() + timeout
        chunks: List[bytes] = []
        while time.time() < deadline:
            try:
                part = self.tn.read_very_eager()
            except EOFError:
                break
            except OSError as exc:
                raise TelnetSessionError(f"讀取設備回應失敗：{cmd} | 原因：{exc}") from exc
            if part:
                chunks.append(part)
                merged = b"".join(chunks)
                if expect_tokens and any(token in merged for token in expect_tokens):
                    break
            time.sleep(READ_INTERVAL)

        if settle > 0:
            time.sleep(settle)
            idle = 0
            while idle < 3:
                try:
                    part = self.tn.read_very_eager()
                except EOFError:
                    break
                except OSError:
                    break
                if part:
                    chunks.append(part)
                    idle = 0
                else:
                    idle += 1
                time.sleep(READ_INTERVAL)

        response = b"".join(chunks).decode(ENCODING, errors="ignore")
        if expect_tokens:
            merged = b"".join(chunks)
            if not any(token in merged for token in expect_tokens):
                raise CommandTimeoutError(f"指令等待逾時：{cmd} | 未等到預期 prompt。最後回應：{sanitize_excel_text(response)[:300]}")
        logger.info("Response for [%s]: %s", cmd, sanitize_excel_text(response).replace("\n", "\\n")[:700])
        return response

    def login(self) -> str:
        outputs = [
            "=== INITIAL BANNER ===", self._read_banner(),
            "=== LOGIN USER ===", self._send_line_and_wait(USERNAME, 8.0, 0.2, [b"Password:"]),
            "=== LOGIN PASSWORD ===", self._send_line_and_wait(PASSWORD, 8.0, 0.2, [b">", b"#"]),
        ]
        return "\n".join(outputs)

    def read_labelinfo(self) -> str:
        outputs = [
            "=== LABELINFO ===",
            self._send_line_and_wait(LABELINFO_CMD, 8.0, 0.3, [b">", b"#", b"SN", b"MAC", b"Version"]),
        ]
        return "\n".join(outputs)


def create_workbook(path: Path) -> None:
    wb = Workbook()
    headers_all = [
        "Time", "Session_ID", "Work_Order", "Inspector", "Model", "Station", "Remark",
        "SN_Label", "MAC_Label", "GPON_SN_Label", "QR_Raw", "QR_SN", "QR_MAC", "QR_WPA",
        "Version_DUT", "SN_DUT", "MAC_DUT", "GPON_SN_DUT",
        "SSID_5G", "WPA_5G", "SSID_24G", "WPA_24G", "MLO_SSID", "MLO_WPA",
        "SN_Match", "MAC_Match", "GPON_SN_Match", "Version_Auto_Check",
        "Auto_Result", "Auto_Fail_Reason",
        "Version_Manual", "SSID_5G_Manual", "WPA_5G_Manual",
        "SSID_24G_Manual", "WPA_24G_Manual", "MLO_SSID_Manual", "MLO_WPA_Manual",
        "Manual_Result", "Manual_Fail_Reason",
        "Final_Result", "Fail_Reason", "Raw_LabelInfo"
    ]
    ws_all = wb.active
    ws_all.title = "ALL"
    ws_all.append(headers_all)

    ws_manual = wb.create_sheet("MANUAL_CONFIRM")
    ws_manual.append([
        "Session_ID", "version_manual", "ssid_5g_manual", "wpa_5g_manual",
        "ssid_24g_manual", "wpa_24g_manual", "mlo_ssid_manual", "mlo_wpa_manual", "remark"
    ])

    ws_pass = wb.create_sheet("PASS")
    ws_pass.append(headers_all)
    ws_fail = wb.create_sheet("FAIL")
    ws_fail.append(headers_all)
    wb.save(path)
    wb.close()


def ensure_workbook(path: Path) -> None:
    if not path.exists():
        create_workbook(path)


def backup_excel(path: Path) -> None:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUP_DIR / f"{path.stem}_{timestamp}{path.suffix}"
    shutil.copy2(path, backup_path)
    logger.info("Backup created: %s", backup_path)


def append_record(path: Path, record: Dict[str, str]) -> None:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws_all = wb["ALL"]
    headers = [cell.value for cell in ws_all[1]]
    row = [sanitize_excel_text(record.get(h, "")) for h in headers]
    ws_all.append(row)

    ws_manual = wb["MANUAL_CONFIRM"]
    ws_manual.append([
        sanitize_excel_text(record.get("Session_ID", "")),
        sanitize_excel_text(record.get("Version_Manual", "")),
        sanitize_excel_text(record.get("SSID_5G_Manual", "")),
        sanitize_excel_text(record.get("WPA_5G_Manual", "")),
        sanitize_excel_text(record.get("SSID_24G_Manual", "")),
        sanitize_excel_text(record.get("WPA_24G_Manual", "")),
        sanitize_excel_text(record.get("MLO_SSID_Manual", "")),
        sanitize_excel_text(record.get("MLO_WPA_Manual", "")),
        sanitize_excel_text(record.get("Manual_Fail_Reason", "")),
    ])

    target_sheet = "PASS" if record.get("Final_Result") == "PASS" else "FAIL"
    wb[target_sheet].append(row)
    wb.save(path)
    wb.close()
    backup_excel(path)
    logger.info("Excel record written: %s | WorkOrder=%s | SN=%s | Final=%s", path, record.get("Work_Order"), record.get("SN_Label"), record.get("Final_Result"))


def count_results(path: Path) -> Tuple[int, int, int]:
    if not path.exists():
        return 0, 0, 0
    wb = load_workbook(path, read_only=True)
    ws = wb["ALL"]
    total = passed = failed = 0
    try:
        header = [c.value for c in ws[1]]
        result_idx = header.index("Final_Result")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            total += 1
            result = str(row[result_idx]).strip().upper() if row[result_idx] else ""
            if result == "PASS":
                passed += 1
            elif result == "FAIL":
                failed += 1
    finally:
        wb.close()
    return total, passed, failed


def run_phase1_check(host: str) -> Dict[str, object]:
    client = GRG4366TelnetClient(host=host)
    login_response = ""
    labelinfo_response = ""
    try:
        client.connect()
        login_response = client.login()
        labelinfo_response = client.read_labelinfo()
        parsed = parse_labelinfo(labelinfo_response)
        raw = f"{login_response}\n{labelinfo_response}"
        parsed["raw_text"] = raw
        return {"ok": True, "raw_text": raw, "parsed": parsed}
    except Exception as exc:
        logger.exception("Phase1 check failed")
        return {"ok": False, "error": str(exc), "raw_text": f"{login_response}\n{labelinfo_response}".strip()}
    finally:
        client.close()


class Phase1InspectionApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(f"{APP_TITLE} {APP_VERSION}")
        self.root.geometry("1580x920")
        self.root.minsize(1450, 860)

        self.host = HOST
        self.excel_file = EXCEL_FILE
        self.result_queue: "queue.Queue[Dict[str, object]]" = queue.Queue()
        self.busy = False
        self.telnet_loaded = False
        self.saved_once = False

        self.session_id = ""
        self.current_dut_data: Dict[str, str] = {}
        self.current_compare: Dict[str, str] = {}
        self.current_qr: Dict[str, str] = {}

        self._init_vars()
        self._build_ui()
        self._load_counts()
        self._poll_queue()
        self._new_session()

    def _init_vars(self) -> None:
        self.work_order_var = tk.StringVar()
        self.inspector_var = tk.StringVar()
        self.model_var = tk.StringVar(value="GRG-4366")
        self.station_var = tk.StringVar()
        self.time_var = tk.StringVar()
        self.expected_version_var = tk.StringVar()
        self.host_var = tk.StringVar(value=self.host)

        self.sn_label_var = tk.StringVar()
        self.mac_label_var = tk.StringVar()
        self.gpon_sn_label_var = tk.StringVar()
        self.qr_raw_var = tk.StringVar()
        self.qr_sn_var = tk.StringVar()
        self.qr_mac_var = tk.StringVar()
        self.qr_wpa_var = tk.StringVar()

        self.status_var = tk.StringVar(value="等待建立檢驗單")
        self.auto_result_var = tk.StringVar(value="-")
        self.manual_result_var = tk.StringVar(value="-")
        self.final_result_var = tk.StringVar(value="-")
        self.fail_reason_var = tk.StringVar(value="")
        self.counters_var = tk.StringVar(value="Total: 0 | PASS: 0 | FAIL: 0")

        self.manual_vars = {
            "version_manual": tk.StringVar(),
            "ssid_5g_manual": tk.StringVar(),
            "wpa_5g_manual": tk.StringVar(),
            "ssid_24g_manual": tk.StringVar(),
            "wpa_24g_manual": tk.StringVar(),
            "mlo_ssid_manual": tk.StringVar(),
            "mlo_wpa_manual": tk.StringVar(),
        }

        self.dut_vars = {
            "version": tk.StringVar(),
            "sn_dut": tk.StringVar(),
            "mac_dut": tk.StringVar(),
            "gpon_sn_dut": tk.StringVar(),
            "ssid_5g": tk.StringVar(),
            "wpa_5g": tk.StringVar(),
            "ssid_24g": tk.StringVar(),
            "wpa_24g": tk.StringVar(),
            "mlo_ssid": tk.StringVar(),
            "mlo_wpa": tk.StringVar(),
        }
        self.compare_vars = {
            "sn_match": tk.StringVar(value="-"),
            "mac_match": tk.StringVar(value="-"),
            "gpon_sn_match": tk.StringVar(value="-"),
            "version_auto_check": tk.StringVar(value="-"),
        }

    def _build_ui(self) -> None:
        main = tk.Frame(self.root, padx=10, pady=8)
        main.pack(fill="both", expand=True)

        header = tk.Frame(main)
        header.pack(fill="x", pady=(0, 8))
        tk.Label(header, text=f"{APP_TITLE}  {APP_VERSION}", font=("Arial", 20, "bold")).pack(side="left")
        tk.Label(header, textvariable=self.counters_var, font=("Arial", 12, "bold")).pack(side="right")

        body = tk.Frame(main)
        body.pack(fill="both", expand=True)

        self.left_nav = tk.Frame(body, width=170, bd=1, relief="groove", padx=8, pady=8)
        self.left_nav.pack(side="left", fill="y", padx=(0, 8))
        self.left_nav.pack_propagate(False)

        self.center = tk.Frame(body)
        self.center.pack(side="left", fill="both", expand=True, padx=(0, 8))

        self.right = tk.Frame(body, width=360, bd=1, relief="groove", padx=8, pady=8)
        self.right.pack(side="right", fill="y")
        self.right.pack_propagate(False)

        self._build_left_nav()
        self._build_center()
        self._build_right()

        bottom = tk.LabelFrame(main, text="操作 Log", padx=6, pady=6)
        bottom.pack(fill="both", expand=False, pady=(8, 0))
        self.log_text = tk.Text(bottom, height=10, font=("Consolas", 10))
        self.log_text.pack(fill="both", expand=True)
        self.log_text.insert("end", "程式啟動完成。\n")
        self.log_text.configure(state="disabled")

    def _build_left_nav(self) -> None:
        tk.Label(self.left_nav, text="流程導航", font=("Arial", 13, "bold")).pack(anchor="w", pady=(0, 8))
        self.step_labels = []
        for text in [
            "01 基本資料",
            "02 Label 輸入",
            "03 Telnet 讀取",
            "04 比對 / 人工確認",
            "05 匯出 / 結案",
        ]:
            lbl = tk.Label(self.left_nav, text=text, anchor="w", justify="left", bg="#f0f0f0", width=18, pady=6)
            lbl.pack(fill="x", pady=2)
            self.step_labels.append(lbl)

        self.current_step = 0
        self._set_step(0)

    def _set_step(self, index: int) -> None:
        self.current_step = index
        for i, lbl in enumerate(self.step_labels):
            if i == index:
                lbl.configure(bg="#9dc3e6", font=("Arial", 10, "bold"))
            else:
                lbl.configure(bg="#f0f0f0", font=("Arial", 10))

    def _build_center(self) -> None:
        top = tk.LabelFrame(self.center, text="基本資料 / Label 輸入", padx=8, pady=8)
        top.pack(fill="x")

        row = 0
        fields = [
            ("工單號", self.work_order_var, 0, 0),
            ("檢驗員", self.inspector_var, 0, 2),
            ("機種", self.model_var, 1, 0),
            ("站別", self.station_var, 1, 2),
            ("檢驗時間", self.time_var, 2, 0),
            ("預期 Version", self.expected_version_var, 2, 2),
            ("SN", self.sn_label_var, 3, 0),
            ("MAC", self.mac_label_var, 3, 2),
            ("GPON SN", self.gpon_sn_label_var, 4, 0),
            ("QR Code", self.qr_raw_var, 4, 2),
        ]
        for label, var, r, c in fields:
            tk.Label(top, text=label).grid(row=r, column=c, sticky="e", padx=5, pady=4)
            width = 42 if label == "QR Code" else 28
            state = "readonly" if label in ("機種", "檢驗時間") else "normal"
            ent = tk.Entry(top, textvariable=var, width=width)
            if state == "readonly":
                ent.configure(state="readonly")
            ent.grid(row=r, column=c + 1, sticky="w", padx=5, pady=4)

        tk.Label(top, text="QR 解析 SN").grid(row=5, column=0, sticky="e", padx=5, pady=4)
        tk.Entry(top, textvariable=self.qr_sn_var, width=28, state="readonly").grid(row=5, column=1, sticky="w", padx=5, pady=4)
        tk.Label(top, text="QR 解析 MAC").grid(row=5, column=2, sticky="e", padx=5, pady=4)
        tk.Entry(top, textvariable=self.qr_mac_var, width=28, state="readonly").grid(row=5, column=3, sticky="w", padx=5, pady=4)
        tk.Label(top, text="QR 解析 WPA").grid(row=6, column=0, sticky="e", padx=5, pady=4)
        tk.Entry(top, textvariable=self.qr_wpa_var, width=28, state="readonly").grid(row=6, column=1, sticky="w", padx=5, pady=4)
        tk.Label(top, text="狀態").grid(row=6, column=2, sticky="e", padx=5, pady=4)
        self.status_entry = tk.Entry(top, textvariable=self.status_var, width=42, state="readonly")
        self.status_entry.grid(row=6, column=3, sticky="w", padx=5, pady=4)

        btns = tk.Frame(top)
        btns.grid(row=7, column=0, columnspan=4, sticky="w", pady=(8, 0))
        tk.Button(btns, text="新建檢驗單", width=14, command=self._new_session).pack(side="left", padx=4)
        tk.Button(btns, text="解析 QR", width=12, command=self._parse_qr_into_fields).pack(side="left", padx=4)
        tk.Button(btns, text="讀取 labelinfo", width=16, command=self._start_telnet_read).pack(side="left", padx=4)
        tk.Button(btns, text="匯出 Excel", width=12, command=self._export_excel).pack(side="left", padx=4)
        tk.Button(btns, text="結案", width=10, command=self._close_case).pack(side="left", padx=4)
        tk.Button(btns, text="開新單", width=10, command=self._new_session).pack(side="left", padx=4)

        middle = tk.PanedWindow(self.center, orient="horizontal", sashrelief="raised")
        middle.pack(fill="both", expand=True, pady=(8, 0))

        raw_frame = tk.LabelFrame(middle, text="labelinfo 原始回傳", padx=6, pady=6)
        self.raw_text = tk.Text(raw_frame, font=("Consolas", 10), width=70)
        self.raw_text.pack(fill="both", expand=True)
        middle.add(raw_frame, minsize=460)

        parsed_frame = tk.LabelFrame(middle, text="DUT 解析欄位", padx=6, pady=6)
        self.parsed_tree = ttk.Treeview(parsed_frame, columns=("field", "value"), show="headings", height=16)
        self.parsed_tree.heading("field", text="欄位")
        self.parsed_tree.heading("value", text="值")
        self.parsed_tree.column("field", width=160, anchor="w")
        self.parsed_tree.column("value", width=320, anchor="w")
        self.parsed_tree.pack(fill="both", expand=True)
        middle.add(parsed_frame, minsize=450)

    def _build_right(self) -> None:
        info = tk.LabelFrame(self.right, text="摘要 / 比對 / 人工確認", padx=8, pady=8)
        info.pack(fill="both", expand=True)

        tk.Label(info, text="DUT IP").grid(row=0, column=0, sticky="e", padx=4, pady=3)
        tk.Entry(info, textvariable=self.host_var, width=22).grid(row=0, column=1, sticky="w", padx=4, pady=3)

        tk.Label(info, text="Auto Result").grid(row=1, column=0, sticky="e", padx=4, pady=3)
        tk.Entry(info, textvariable=self.auto_result_var, width=22, state="readonly").grid(row=1, column=1, sticky="w", padx=4, pady=3)
        tk.Label(info, text="Manual Result").grid(row=2, column=0, sticky="e", padx=4, pady=3)
        tk.Entry(info, textvariable=self.manual_result_var, width=22, state="readonly").grid(row=2, column=1, sticky="w", padx=4, pady=3)
        tk.Label(info, text="Final Result").grid(row=3, column=0, sticky="e", padx=4, pady=3)
        tk.Entry(info, textvariable=self.final_result_var, width=22, state="readonly").grid(row=3, column=1, sticky="w", padx=4, pady=3)

        compare_box = tk.LabelFrame(info, text="自動比對", padx=6, pady=6)
        compare_box.grid(row=4, column=0, columnspan=2, sticky="we", pady=(8, 6))
        crow = 0
        for k, title in [
            ("sn_match", "SN"),
            ("mac_match", "MAC"),
            ("gpon_sn_match", "GPON SN"),
            ("version_auto_check", "Version"),
        ]:
            tk.Label(compare_box, text=title, width=12, anchor="e").grid(row=crow, column=0, sticky="e", padx=3, pady=2)
            ent = tk.Entry(compare_box, textvariable=self.compare_vars[k], width=12, justify="center", state="readonly")
            ent.grid(row=crow, column=1, sticky="w", padx=3, pady=2)
            crow += 1

        manual_box = tk.LabelFrame(info, text="人工確認", padx=6, pady=6)
        manual_box.grid(row=5, column=0, columnspan=2, sticky="we", pady=(4, 6))
        mrow = 0
        manual_titles = [
            ("version_manual", "Version"),
            ("ssid_5g_manual", "5G SSID"),
            ("wpa_5g_manual", "5G WPA"),
            ("ssid_24g_manual", "2.4G SSID"),
            ("wpa_24g_manual", "2.4G WPA"),
            ("mlo_ssid_manual", "MLO SSID"),
            ("mlo_wpa_manual", "MLO WPA"),
        ]
        for key, title in manual_titles:
            tk.Label(manual_box, text=title, width=12, anchor="e").grid(row=mrow, column=0, sticky="e", padx=3, pady=2)
            box = ttk.Combobox(manual_box, textvariable=self.manual_vars[key], values=RESULT_OPTIONS, width=10, state="readonly")
            box.grid(row=mrow, column=1, sticky="w", padx=3, pady=2)
            box.bind("<<ComboboxSelected>>", lambda e: self._refresh_results())
            mrow += 1

        tk.Label(info, text="異常原因 / 備註").grid(row=6, column=0, columnspan=2, sticky="w", padx=4)
        self.manual_reason_text = tk.Text(info, height=8, width=34, font=("Arial", 10))
        self.manual_reason_text.grid(row=7, column=0, columnspan=2, sticky="we", padx=4, pady=(2, 6))
        self.manual_reason_text.bind("<KeyRelease>", lambda e: self._refresh_results())

        info.grid_columnconfigure(1, weight=1)

    def _append_ui_log(self, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _set_status(self, text: str) -> None:
        self.status_var.set(text)
        self._append_ui_log(text)

    def _load_counts(self) -> None:
        total, passed, failed = count_results(self.excel_file)
        self.counters_var.set(f"Total: {total} | PASS: {passed} | FAIL: {failed}")

    def _parse_qr_into_fields(self) -> None:
        parsed = parse_qrcode(self.qr_raw_var.get())
        self.current_qr = parsed
        self.qr_sn_var.set(parsed.get("qr_sn", ""))
        self.qr_mac_var.set(parsed.get("qr_mac", ""))
        self.qr_wpa_var.set(parsed.get("qr_wpa", ""))

        if not self.sn_label_var.get().strip() and parsed.get("qr_sn"):
            self.sn_label_var.set(parsed["qr_sn"])
        if not self.mac_label_var.get().strip() and parsed.get("qr_mac"):
            self.mac_label_var.set(parsed["qr_mac"])
        self._set_status("QR Code 已解析")

    def _set_all_manual_result(self, value: str) -> None:
        for var in self.manual_vars.values():
            var.set(value)
        self._refresh_results()
        self._set_status(f"人工確認已批次設定為 {value}")

    def _mark_all_manual_pass(self) -> None:
        self._set_all_manual_result("PASS")

    def _mark_all_manual_na(self) -> None:
        self._set_all_manual_result("N/A")

    def _reset_display(self) -> None:
        self.raw_text.delete("1.0", "end")
        for item in self.parsed_tree.get_children():
            self.parsed_tree.delete(item)
        self.current_dut_data = {}
        self.current_compare = {}
        self.telnet_loaded = False

        for var in self.compare_vars.values():
            var.set("-")
        for var in self.manual_vars.values():
            var.set("")
        self.manual_reason_text.delete("1.0", "end")
        self.auto_result_var.set("-")
        self.manual_result_var.set("-")
        self.final_result_var.set("-")
        self.fail_reason_var.set("")

    def _new_session(self) -> None:
        self.session_id = datetime.now().strftime("%Y%m%d%H%M%S")
        self.time_var.set(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        self.work_order_var.set("")
        self.inspector_var.set("")
        self.station_var.set("")
        self.expected_version_var.set("")
        self.sn_label_var.set("")
        self.mac_label_var.set("")
        self.gpon_sn_label_var.set("")
        self.qr_raw_var.set("")
        self.qr_sn_var.set("")
        self.qr_mac_var.set("")
        self.qr_wpa_var.set("")
        self.current_qr = {}
        self.saved_once = False
        self._reset_display()
        self._set_step(0)
        self.status_var.set("已建立新檢驗單，請輸入基本資料")
        self._append_ui_log(f"新建檢驗單: {self.session_id}")

    def _validate_basic(self) -> bool:
        if not self.work_order_var.get().strip():
            messagebox.showerror("缺少資料", "工單號不可空白")
            self._set_step(0)
            return False
        if not self.inspector_var.get().strip():
            messagebox.showerror("缺少資料", "檢驗員不可空白")
            self._set_step(0)
            return False
        return True

    def _validate_label_input(self) -> bool:
        if not validate_sn(self.sn_label_var.get()):
            messagebox.showerror("格式錯誤", "SN 格式不正確")
            self._set_step(1)
            return False
        if not validate_mac(self.mac_label_var.get()):
            messagebox.showerror("格式錯誤", "MAC 格式不正確")
            self._set_step(1)
            return False
        if not validate_gpon_sn(self.gpon_sn_label_var.get()):
            messagebox.showerror("格式錯誤", "GPON SN 格式不正確")
            self._set_step(1)
            return False
        return True

    def _start_telnet_read(self) -> None:
        if self.busy:
            return
        if not self._validate_basic():
            return
        if not self._validate_label_input():
            return

        self._parse_qr_into_fields()
        self._set_step(2)
        self._set_status("Telnet 讀取中...")
        self.busy = True
        thread = threading.Thread(target=self._worker_telnet, daemon=True)
        thread.start()

    def _worker_telnet(self) -> None:
        host = self.host_var.get().strip() or HOST
        result = run_phase1_check(host=host)
        self.result_queue.put(result)

    def _poll_queue(self) -> None:
        try:
            result = self.result_queue.get_nowait()
        except queue.Empty:
            self.root.after(120, self._poll_queue)
            return

        self.busy = False
        if result.get("ok"):
            self.current_dut_data = result["parsed"]
            self.telnet_loaded = True
            self._fill_raw_and_parsed(result["raw_text"], self.current_dut_data)
            self._do_compare()
            self._set_step(3)
            self._set_status("labelinfo 讀取成功，請確認比對與人工查核")
        else:
            self.telnet_loaded = False
            self.raw_text.delete("1.0", "end")
            self.raw_text.insert("1.0", sanitize_excel_text(result.get("raw_text", "")))
            messagebox.showerror("Telnet 失敗", sanitize_excel_text(result.get("error", "未知錯誤")))
            self._set_status("Telnet 讀取失敗")
        self.root.after(120, self._poll_queue)

    def _fill_raw_and_parsed(self, raw_text: str, parsed: Dict[str, str]) -> None:
        self.raw_text.delete("1.0", "end")
        self.raw_text.insert("1.0", sanitize_excel_text(raw_text))
        for item in self.parsed_tree.get_children():
            self.parsed_tree.delete(item)

        order = [
            ("version", "Version"),
            ("sn_dut", "SN"),
            ("mac_dut", "MAC"),
            ("gpon_sn_dut", "GPON_SN"),
            ("ssid_5g", "5G SSID"),
            ("wpa_5g", "5G WPAKey"),
            ("ssid_24g", "2.4G SSID"),
            ("wpa_24g", "2.4G WPAKey"),
            ("mlo_ssid", "MLO SSID"),
            ("mlo_wpa", "MLO WPAKey"),
        ]
        for key, title in order:
            value = sanitize_excel_text(parsed.get(key, ""))
            self.parsed_tree.insert("", "end", values=(title, value))
            if key in self.dut_vars:
                self.dut_vars[key].set(value)

    def _label_data(self) -> Dict[str, str]:
        qr = self.current_qr if self.current_qr else parse_qrcode(self.qr_raw_var.get())
        return {
            "sn_label": normalize_sn(self.sn_label_var.get()),
            "mac_label": normalize_mac(self.mac_label_var.get()),
            "gpon_sn_label": normalize_gpon_sn(self.gpon_sn_label_var.get()),
            "qr_raw": sanitize_excel_text(self.qr_raw_var.get()),
            "qr_sn": qr.get("qr_sn", ""),
            "qr_mac": qr.get("qr_mac", ""),
            "qr_wpa": qr.get("qr_wpa", ""),
        }

    def _do_compare(self) -> None:
        self.current_compare = compare_fields(
            self._label_data(),
            self.current_dut_data,
            self.expected_version_var.get(),
        )
        for key in self.compare_vars:
            self.compare_vars[key].set(self.current_compare.get(key, "-"))
        self.auto_result_var.set(self.current_compare.get("auto_result", "-"))
        self._refresh_results()

    def _manual_data(self) -> Dict[str, str]:
        data = {k: v.get().strip() for k, v in self.manual_vars.items()}
        data["manual_fail_reason"] = sanitize_excel_text(self.manual_reason_text.get("1.0", "end"))
        result_list = [data[k] for k in self.manual_vars]
        if any(v == "FAIL" for v in result_list):
            data["manual_result"] = "FAIL"
        elif all(v in ("PASS", "N/A") for v in result_list):
            data["manual_result"] = "PASS"
        else:
            data["manual_result"] = "PENDING"
        return data

    def _compute_final(self) -> Tuple[str, str]:
        compare_data = self.current_compare
        manual_data = self._manual_data()

        reasons = []
        if compare_data.get("auto_result") == "FAIL":
            reasons.append(compare_data.get("auto_fail_reason", "").strip())
        if manual_data.get("manual_result") == "FAIL":
            manual_reason = manual_data.get("manual_fail_reason", "").strip() or "人工確認項目 FAIL"
            reasons.append(manual_reason)

        if compare_data.get("auto_result") == "FAIL" or manual_data.get("manual_result") == "FAIL":
            return "FAIL", "；".join([r for r in reasons if r])
        if not self.telnet_loaded:
            return "PENDING", "尚未讀取 labelinfo"
        if manual_data.get("manual_result") != "PASS":
            return "PENDING", "人工確認未完成"
        return "PASS", ""

    def _refresh_results(self) -> None:
        manual = self._manual_data()
        self.manual_result_var.set(manual.get("manual_result", "-"))
        final, reason = self._compute_final()
        self.final_result_var.set(final)
        self.fail_reason_var.set(reason)

    def _build_export_record(self) -> Dict[str, str]:
        if not self.telnet_loaded:
            raise ValueError("尚未讀取 labelinfo")
        compare_data = self.current_compare
        manual_data = self._manual_data()
        final_result, fail_reason = self._compute_final()
        if final_result == "PENDING":
            raise ValueError(fail_reason or "尚有未完成項目")

        return build_record(
            work_order=self.work_order_var.get().strip(),
            inspector=self.inspector_var.get().strip(),
            model_name=self.model_var.get().strip(),
            station_id=self.station_var.get().strip(),
            remark=sanitize_excel_text(self.manual_reason_text.get("1.0", "end")),
            label_data=self._label_data(),
            dut_data=self.current_dut_data,
            compare_data=compare_data,
            manual_data=manual_data,
            final_result=final_result,
            fail_reason=fail_reason,
            session_id=self.session_id,
        )

    def _export_excel(self) -> None:
        try:
            if not self._validate_basic():
                return
            if not self._validate_label_input():
                return
            record = self._build_export_record()
        except Exception as exc:
            messagebox.showerror("無法匯出", str(exc))
            return
        append_record(self.excel_file, record)
        self.saved_once = True
        self._set_step(4)
        self._load_counts()
        self._set_status(f"Excel 已匯出：{self.excel_file.name}")
        messagebox.showinfo("完成", f"已匯出到：\n{self.excel_file}")

    def _close_case(self) -> None:
        try:
            record = self._build_export_record()
        except Exception as exc:
            messagebox.showerror("無法結案", str(exc))
            return

        if not self.saved_once:
            if messagebox.askyesno("尚未匯出", "尚未匯出 Excel，是否先自動匯出再結案？"):
                append_record(self.excel_file, record)
                self.saved_once = True
                self._load_counts()
            else:
                return

        self._set_step(4)
        self._set_status(f"檢驗單已結案 | Final Result = {record['Final_Result']}")
        messagebox.showinfo("結案完成", f"Session {self.session_id} 已結案。\n結果：{record['Final_Result']}")

    def run(self) -> None:
        self.root.mainloop()


def main() -> None:
    logger.info("Application started")
    root = tk.Tk()
    app = Phase1InspectionApp(root)
    app.run()


if __name__ == "__main__":
    main()
