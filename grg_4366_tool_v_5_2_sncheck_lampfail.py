import logging
import queue
import re
import shutil
import socket
import telnetlib
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Tuple, List

import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

APP_VERSION = "V5.3"
HOST = "192.168.1.1"
PORT = 23
TIMEOUT = 12
ENCODING = "ascii"
USERNAME = "root"
PASSWORD = "12345"
DEBUG_CMD = "version --debug x4Wnhq2ReL"
SERIAL_CMD = "serialnumber"

CONFIG_COMMAND_FLOW = [
    {"cmd": "sh", "timeout": 8.0, "settle": 0.2, "expect_tokens": [b"#"]},
    {"cmd": "app-cli", "timeout": 15.0, "settle": 0.5, "expect_tokens": [b"Cortina>", b"Cortina#"]},
    {"cmd": "enable", "timeout": 8.0, "settle": 0.2, "expect_tokens": [b"Cortina#"]},
    {"cmd": "config", "timeout": 8.0, "settle": 0.2, "expect_tokens": [b"Cortina(config)#", b"(config)#"]},
    {"cmd": "aal", "timeout": 8.0, "settle": 0.2, "expect_tokens": [b"Cortina(config-aal)#", b"(config-aal)#"]},
    {"cmd": "xgpon", "timeout": 8.0, "settle": 0.2, "expect_tokens": [b"Cortina(config-aal-xgpon)#", b"(config-aal-xgpon)#"]},
    {"cmd": "bwmp_rsp_tm 0 0x1543", "timeout": 10.0, "settle": 0.3, "expect_tokens": [b"Cortina(config-aal-xgpon)#", b"#", b">"]},
]

INITIAL_READ_DELAY = 1.0
READ_INTERVAL = 0.15
READ_IDLE_BREAK = 4
SUCCESS_KEYWORDS = ["ok", "success", "done", "complete", "set"]
FAIL_KEYWORDS = ["error", "fail", "invalid", "unknown", "denied", "not found", "incorrect"]

APP_TITLE = "GRG-4366 產線設定工具"
BASE_DIR = Path.cwd()
LOG_DIR = BASE_DIR / "logs"
OUTPUT_DIR = BASE_DIR / "output"
BACKUP_DIR = BASE_DIR / "backup"

QR_PARSE_PATTERNS = {
    "sn": [
        r"(?i)S\s*/\s*N\s*[:：]\s*([A-Za-z0-9\-_]+)",
        r"(?i)SN\s*[:：]\s*([A-Za-z0-9\-_]+)",
        r"(?i)Serial\s*(?:Number)?\s*[:：]\s*([A-Za-z0-9\-_]+)",
    ],
    "mac": [
        r"(?i)MAC\s*[:：]\s*([0-9A-F]{12})",
        r"(?i)MAC\s*[:：]\s*([0-9A-F]{2}(?::[0-9A-F]{2}){5})",
        r"(?i)MAC\s*[:：]\s*([0-9A-F]{2}(?:-[0-9A-F]{2}){5})",
    ],
    "wpa": [
        r"(?i)WPA(?:\s*Key)?\s*[:：]\s*([^\s]+)",
        r"(?i)WiFi(?:\s*Key)?\s*[:：]\s*([^\s]+)",
    ],
}

class TelnetSessionError(Exception):
    pass

class CommandTimeoutError(TelnetSessionError):
    pass

def ensure_dirs() -> None:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

def get_daily_paths() -> Tuple[Path, Path]:
    date_tag = datetime.now().strftime("%Y%m%d")
    excel_path = OUTPUT_DIR / f"grg_4366_setting_log_{date_tag}.xlsx"
    log_path = LOG_DIR / f"grg_4366_setting_log_{date_tag}.log"
    return excel_path, log_path

ensure_dirs()
EXCEL_FILE, LOG_FILE = get_daily_paths()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_FILE, encoding="utf-8"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

def sanitize_excel_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("
", "
").replace("
", "
")
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text

def normalize_mac(mac: str) -> str:
    value = re.sub(r"[^0-9A-Fa-f]", "", mac or "").upper()
    return value[:12] if len(value) >= 12 else value

def parse_qr_payload(payload: str) -> Dict[str, str]:
    text = sanitize_excel_text(payload).strip()
    compact = re.sub(r"\s+", "", text)
    result = {"sn": "", "mac": "", "wpa": ""}
    for key, patterns in QR_PARSE_PATTERNS.items():
        for pattern in patterns:
            match = re.search(pattern, text) or re.search(pattern, compact)
            if match:
                result[key] = match.group(1).strip()
                break
    if result["mac"]:
        result["mac"] = normalize_mac(result["mac"])
    return result

def parse_device_sn(response: str, input_sn: str = "") -> str:
    text = sanitize_excel_text(response)
    if input_sn and input_sn in text:
        return input_sn
    patterns = [
        r"(?i)serial\s*number\s*[:=]\s*([A-Za-z0-9\-_]+)",
        r"(?i)serialnumber\s*[:=]?\s*([A-Za-z0-9\-_]+)",
        r"(?i)SN\s*[:=]\s*([A-Za-z0-9\-_]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1).strip()
    tokens = re.findall(r"[A-Za-z0-9][A-Za-z0-9\-_]{7,}", text)
    blacklist = {"serialnumber", "serial", "number", "cortina", "password", "login", "successfully", "warning", "config", "enable", "debug"}
    candidates = []
    for token in tokens:
        if token.lower() in blacklist or token.isdigit():
            continue
        if len(token) >= 10:
            candidates.append(token)
    candidates.sort(key=lambda s: (0 if "-" in s or "_" in s else 1, -len(s)))
    return candidates[0] if candidates else ""

def build_record(*, input_sn: str, device_sn: str, sn_check_result: str, host: str,
                 config_result: str, lamp_check_result: str, final_result: str,
                 matched: str, raw_response: str, note: str,
                 qr_raw: str = "", qr_sn: str = "", qr_mac: str = "", qr_wpa_key: str = "",
                 time_text: Optional[str] = None) -> Dict[str, str]:
    return {
        "time": time_text or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "qr_raw": qr_raw,
        "qr_sn": qr_sn,
        "qr_mac": qr_mac,
        "qr_wpa_key": qr_wpa_key,
        "input_sn": input_sn,
        "device_sn": device_sn,
        "sn_check_result": sn_check_result,
        "host": host,
        "config_result": config_result,
        "lamp_check_result": lamp_check_result,
        "final_result": final_result,
        "matched": matched,
        "raw_response": raw_response,
        "note": note,
    }

class GRG4366Configurator:
    def __init__(self, host: str = HOST, port: int = PORT, timeout: int = TIMEOUT):
        self.host = host
        self.port = port
        self.timeout = timeout
        self.tn: Optional[telnetlib.Telnet] = None

    def connect(self) -> None:
        try:
            logger.info("Connecting to %s:%s", self.host, self.port)
            self.tn = telnetlib.Telnet(self.host, self.port, self.timeout)
        except Exception as exc:
            raise TelnetSessionError(f"無法連線到 {self.host}:{self.port}，原因：{exc}") from exc

    def close(self) -> None:
        if self.tn:
            try:
                self.tn.close()
            except Exception:
                pass
            self.tn = None

    def _read_banner(self) -> str:
        if not self.tn:
            raise TelnetSessionError("Telnet 尚未連線")
        time.sleep(INITIAL_READ_DELAY)
        chunks: List[bytes] = []
        idle = 0
        while idle < READ_IDLE_BREAK:
            try:
                part = self.tn.read_very_eager()
            except EOFError:
                break
            except OSError as exc:
                raise TelnetSessionError(f"讀取初始畫面失敗：{exc}") from exc
            if part:
                chunks.append(part)
                idle = 0
            else:
                idle += 1
            time.sleep(READ_INTERVAL)
        return b"".join(chunks).decode(ENCODING, errors="ignore")

    def _send_line_and_wait(self, cmd: str, timeout: float, settle: float = 0.0,
                            expect_tokens: Optional[List[bytes]] = None) -> str:
        if not self.tn:
            raise TelnetSessionError("Telnet 尚未連線")
        expect_tokens = expect_tokens or []
        logger.info("Send command: %s", cmd)
        try:
            self.tn.write((cmd + "
").encode(ENCODING))
        except (EOFError, BrokenPipeError, ConnectionResetError, ConnectionAbortedError, socket.error, OSError) as exc:
            raise TelnetSessionError(f"送出指令時連線中斷：{cmd} | 原因：{exc}") from exc
        deadline = time.time() + timeout
        chunks: List[bytes] = []
        while time.time() < deadline:
            try:
                part = self.tn.read_very_eager()
            except EOFError:
                break
            except OSError as exc:
                raise TelnetSessionError(f"讀取設備回應失敗：{cmd} | 原因：{exc}") from exc
            if part:
                chunks.append(part)
                merged = b"".join(chunks)
                if expect_tokens and any(token in merged for token in expect_tokens):
                    break
            time.sleep(READ_INTERVAL)
        if settle > 0:
            time.sleep(settle)
        idle = 0
        while idle < 3:
            try:
                part = self.tn.read_very_eager()
            except (EOFError, OSError):
                break
            if part:
                chunks.append(part)
                idle = 0
            else:
                idle += 1
            time.sleep(READ_INTERVAL)
        response = b"".join(chunks).decode(ENCODING, errors="ignore")
        logger.info("Response for [%s]: %s", cmd, sanitize_excel_text(response).replace("
", "\n")[:700])
        if expect_tokens:
            merged = b"".join(chunks)
            if not any(token in merged for token in expect_tokens):
                raise CommandTimeoutError(f"指令等待逾時：{cmd} | 未等到預期 prompt。最後回應：{sanitize_excel_text(response)[:300]}")
        return response

    def login(self) -> str:
        outputs = [
            "=== INITIAL BANNER ===", self._read_banner(),
            "=== LOGIN USER ===", self._send_line_and_wait(USERNAME, 8.0, 0.2, [b"Password:"]),
            "=== LOGIN PASSWORD ===", self._send_line_and_wait(PASSWORD, 8.0, 0.2, [b">", b"#"]),
        ]
        return "
".join(outputs)

    def read_serialnumber(self) -> str:
        outputs = [
            "=== SERIALNUMBER CHECK ===",
            self._send_line_and_wait(SERIAL_CMD, 8.0, 0.2, [b">", b"#", b"Serial", b"SN", b"serial"]),
        ]
        return "
".join(outputs)

    def configure(self) -> str:
        outputs = [
            "=== DEBUG MODE ===",
            self._send_line_and_wait(DEBUG_CMD, 8.0, 0.3, [b"debug mode successfully", b">", b"#"]),
        ]
        for item in CONFIG_COMMAND_FLOW:
            outputs.append(f"=== COMMAND: {item['cmd']} ===")
            outputs.append(self._send_line_and_wait(item["cmd"], float(item["timeout"]), float(item["settle"]), item["expect_tokens"]))
        return "
".join(outputs)

def create_workbook(path: Path) -> None:
    wb = Workbook()
    headers = ["Time", "QR_Raw", "QR_SN", "QR_MAC", "QR_WPA_Key", "Input_SN", "Device_SN", "SN_Check_Result", "Host", "Config_Result", "Lamp_Check_Result", "Final_Result", "Matched", "Raw_Response", "Note"]
    ws_all = wb.active
    ws_all.title = "ALL"
    ws_all.append(headers)
    ws_pass = wb.create_sheet("PASS")
    ws_pass.append(headers)
    ws_fail = wb.create_sheet("FAIL")
    ws_fail.append(headers)
    wb.save(path)
    wb.close()

def ensure_workbook(path: Path) -> None:
    if not path.exists():
        create_workbook(path)

def backup_excel(path: Path) -> None:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUP_DIR / f"{path.stem}_{timestamp}{path.suffix}"
    shutil.copy2(path, backup_path)
    logger.info("Backup created: %s", backup_path)

def append_record(path: Path, record: Dict[str, str]) -> None:
    ensure_workbook(path)
    wb = load_workbook(path)
    row = [sanitize_excel_text(record[k]) for k in ["time", "qr_raw", "qr_sn", "qr_mac", "qr_wpa_key", "input_sn", "device_sn", "sn_check_result", "host", "config_result", "lamp_check_result", "final_result", "matched", "raw_response", "note"]]
    wb["ALL"].append(row)
    (wb["PASS"] if record["final_result"] == "PASS" else wb["FAIL"]).append(row)
    wb.save(path)
    wb.close()
    backup_excel(path)
    logger.info("Excel record written: %s | Input_SN=%s | Device_SN=%s | Final=%s", path, record["input_sn"], record["device_sn"], record["final_result"])

def validate_sn(sn: str) -> bool:
    sn = sn.strip()
    return bool(sn and re.fullmatch(r"[A-Za-z0-9\-_]+", sn))

def sn_exists(path: Path, sn: str) -> bool:
    if not path.exists():
        return False
    wb = load_workbook(path, read_only=True)
    ws = wb["ALL"]
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 5 and str(row[5]).strip() == sn:
                return True
    finally:
        wb.close()
    return False

def judge_response(resp: str) -> Dict[str, str]:
    text = sanitize_excel_text(resp).lower()
    matched_fail = [k for k in FAIL_KEYWORDS if k in text]
    matched_pass = [k for k in SUCCESS_KEYWORDS if k in text]
    if matched_fail:
        return {"result": "FAIL", "matched": ", ".join(matched_fail)}
    if "bwmp_rsp_tm 0 0x1543" in text or matched_pass:
        return {"result": "PASS", "matched": ", ".join(matched_pass) if matched_pass else "Final command sent"}
    return {"result": "PASS", "matched": "No fail keyword found"}

def run_setting(input_sn: str, host: str = HOST, qr_data: Optional[Dict[str, str]] = None) -> Dict[str, object]:
    qr_data = qr_data or {}
    device = GRG4366Configurator(host=host)
    login_response = ""
    sn_response = ""
    config_response = ""
    try:
        device.connect()
        login_response = device.login()
        sn_response = device.read_serialnumber()
        device_sn = parse_device_sn(sn_response, input_sn=input_sn)
        sn_match = (device_sn == input_sn and device_sn != "")
        logger.info("SN check | Input_SN=%s | Device_SN=%s | Match=%s", input_sn, device_sn, sn_match)
        if not sn_match:
            record = build_record(
                input_sn=input_sn, device_sn=device_sn or "UNKNOWN", sn_check_result="FAIL", host=host,
                config_result="NOT_RUN", lamp_check_result="NOT_RUN", final_result="FAIL",
                matched="SNMismatch", raw_response=f"{login_response}
{sn_response}",
                note="serialnumber 比對失敗，停止下參數",
                qr_raw=qr_data.get("raw", ""), qr_sn=qr_data.get("sn", ""), qr_mac=qr_data.get("mac", ""), qr_wpa_key=qr_data.get("wpa", ""),
            )
            return {"mode": "final", "record": record}
        config_response = device.configure()
        judged = judge_response(config_response)
        pending_record = build_record(
            input_sn=input_sn, device_sn=device_sn, sn_check_result="PASS", host=host,
            config_result=judged["result"], lamp_check_result="PENDING", final_result="PENDING",
            matched=judged["matched"], raw_response=f"{login_response}
{sn_response}
{config_response}",
            note="等待 PON/GPON 燈號人工確認",
            qr_raw=qr_data.get("raw", ""), qr_sn=qr_data.get("sn", ""), qr_mac=qr_data.get("mac", ""), qr_wpa_key=qr_data.get("wpa", ""),
        )
        return {"mode": "wait_lamp", "record": pending_record}
    except Exception as exc:
        logger.exception("Input_SN=%s setting failed", input_sn)
        record = build_record(
            input_sn=input_sn, device_sn=parse_device_sn(sn_response, input_sn="") or "UNKNOWN",
            sn_check_result="FAIL" if not sn_response else "PASS", host=host,
            config_result="FAIL", lamp_check_result="NOT_RUN", final_result="FAIL",
            matched="Exception", raw_response=f"{login_response}
{sn_response}
{config_response}
{exc}",
            note="程式例外中止",
            qr_raw=qr_data.get("raw", ""), qr_sn=qr_data.get("sn", ""), qr_mac=qr_data.get("mac", ""), qr_wpa_key=qr_data.get("wpa", ""),
        )
        return {"mode": "final", "record": record}
    finally:
        device.close()

class ProductionApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(f"{APP_TITLE} {APP_VERSION}")
        self.root.geometry("1180x720")
        self.root.minsize(1120, 700)
        self.host = HOST
        self.excel_file = EXCEL_FILE
        self.pass_count = 0
        self.fail_count = 0
        self.total_count = 0
        self.last_sn = ""
        self.busy = False
        self.await_lamp_confirm = False
        self.pending_record: Optional[Dict[str, str]] = None
        self.current_qr_data = {"raw": "", "sn": "", "mac": "", "wpa": ""}
        self.result_queue = queue.Queue()
        self._build_ui()
        self._refresh_focus()
        self._load_today_counts()
        self._poll_queue()

    def _build_ui(self) -> None:
        top = tk.Frame(self.root, padx=16, pady=12)
        top.pack(fill="x")
        tk.Label(top, text=f"{APP_TITLE} {APP_VERSION}", font=("Arial", 20, "bold")).pack(anchor="w")
        tk.Label(top, text=f"主機 IP: {self.host} Excel: {self.excel_file.name} Log: {LOG_FILE.name}", font=("Arial", 10)).pack(anchor="w", pady=(6, 0))

        mid = tk.Frame(self.root, padx=16, pady=10)
        mid.pack(fill="x")
        tk.Label(mid, text="請掃描 QR 或輸入 SN", font=("Arial", 16, "bold")).grid(row=0, column=0, sticky="w", columnspan=6)

        self.qr_var = tk.StringVar()
        self.sn_var = tk.StringVar()
        self.qr_info_var = tk.StringVar(value="QR_SN: - | QR_MAC: - | QR_WPA: -")

        tk.Label(mid, text="QR 內容", font=("Arial", 11, "bold")).grid(row=1, column=0, sticky="w")
        self.qr_entry = tk.Entry(mid, textvariable=self.qr_var, font=("Consolas", 14), width=72)
        self.qr_entry.grid(row=2, column=0, sticky="w", pady=(4, 8), columnspan=4)
        self.qr_entry.bind("<Return>", self._on_qr_enter)

        tk.Label(mid, text="SN", font=("Arial", 11, "bold")).grid(row=1, column=4, sticky="w", padx=(12, 0))
        self.sn_entry = tk.Entry(mid, textvariable=self.sn_var, font=("Consolas", 18), width=24)
        self.sn_entry.grid(row=2, column=4, sticky="w", padx=(12, 0), pady=(4, 8), columnspan=2)
        self.sn_entry.bind("<Return>", self._on_enter)

        tk.Label(mid, textvariable=self.qr_info_var, font=("Consolas", 10), fg="#1f1f1f").grid(row=3, column=0, columnspan=6, sticky="w")

        btn_bar = tk.Frame(mid)
        btn_bar.grid(row=4, column=0, columnspan=6, sticky="w", pady=(8, 0))
        self.start_btn = tk.Button(btn_bar, text="測試開始", width=14, height=2, font=("Arial", 11, "bold"), command=self._start_test)
        self.start_btn.pack(side="left", padx=(0, 8))
        self.confirm_btn = tk.Button(btn_bar, text="全部 PASS / 換下一台", width=22, height=2, font=("Arial", 11, "bold"), command=self._confirm_next_unit, state="disabled")
        self.confirm_btn.pack(side="left", padx=(0, 8))
        self.lamp_fail_btn = tk.Button(btn_bar, text="全部 N/A 或燈號 FAIL", width=22, height=2, font=("Arial", 11, "bold"), command=self._record_lamp_fail, state="disabled")
        self.lamp_fail_btn.pack(side="left", padx=(0, 8))
        self.exit_btn = tk.Button(btn_bar, text="離開工具", width=12, height=2, font=("Arial", 11, "bold"), command=self._exit_tool)
        self.exit_btn.pack(side="left")

        self.status_box = tk.Label(mid, text="等待掃碼", font=("Arial", 12, "bold"), width=90, height=2, bg="#d9d9d9", fg="black", relief="groove", anchor="w", padx=10)
        self.status_box.grid(row=5, column=0, columnspan=6, sticky="we", pady=(10, 4))

        counters = tk.Frame(self.root, padx=16, pady=8)
        counters.pack(fill="x")
        self.total_var = tk.StringVar(value="Total: 0")
        self.pass_var = tk.StringVar(value="PASS: 0")
        self.fail_var = tk.StringVar(value="FAIL: 0")
        self.last_var = tk.StringVar(value="Last SN: -")
        tk.Label(counters, textvariable=self.total_var, font=("Arial", 14, "bold")).grid(row=0, column=0, padx=(0, 20), sticky="w")
        tk.Label(counters, textvariable=self.pass_var, font=("Arial", 14, "bold")).grid(row=0, column=1, padx=(0, 20), sticky="w")
        tk.Label(counters, textvariable=self.fail_var, font=("Arial", 14, "bold")).grid(row=0, column=2, padx=(0, 20), sticky="w")
        tk.Label(counters, textvariable=self.last_var, font=("Arial", 13)).grid(row=1, column=0, columnspan=3, pady=(10, 0), sticky="w")

        bottom = tk.Frame(self.root, padx=16, pady=16)
        bottom.pack(fill="both", expand=True)
        self.log_text = tk.Text(bottom, height=20, font=("Consolas", 10))
        self.log_text.pack(fill="both", expand=True)
        self.log_text.insert("end", "程式啟動完成，等待掃碼...
")
        self.log_text.configure(state="disabled")

    def _append_ui_log(self, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{timestamp}] {message}
")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _set_status(self, text: str, bg: str, fg: str = "black") -> None:
        self.status_box.configure(text=text, bg=bg, fg=fg)

    def _refresh_focus(self) -> None:
        self.root.after(100, lambda: self.qr_entry.focus_set())

    def _load_today_counts(self) -> None:
        if not self.excel_file.exists():
            self._update_counters()
            return
        wb = load_workbook(self.excel_file, read_only=True)
        ws = wb["ALL"]
        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                self.total_count += 1
                result = str(row[11]).strip().upper() if len(row) > 11 and row[11] is not None else ""
                if result == "PASS":
                    self.pass_count += 1
                else:
                    self.fail_count += 1
                self.last_sn = str(row[5]).strip()
        finally:
            wb.close()
        self._update_counters()

    def _update_counters(self) -> None:
        self.total_var.set(f"Total: {self.total_count}")
        self.pass_var.set(f"PASS: {self.pass_count}")
        self.fail_var.set(f"FAIL: {self.fail_count}")
        self.last_var.set(f"Last SN: {self.last_sn or '-'}")

    def _play_sound(self, success: bool) -> None:
        try:
            import winsound
            if success:
                winsound.Beep(1600, 180)
                winsound.Beep(2000, 220)
            else:
                winsound.Beep(700, 350)
                winsound.Beep(500, 400)
        except Exception:
            pass

    def _set_busy(self, busy: bool) -> None:
        self.busy = busy
        state_entry = "disabled" if busy or self.await_lamp_confirm else "normal"
        state_start = "disabled" if busy or self.await_lamp_confirm else "normal"
        self.qr_entry.configure(state=state_entry)
        self.sn_entry.configure(state=state_entry)
        self.start_btn.configure(state=state_start)

    def _set_wait_buttons(self, enabled: bool) -> None:
        state = "normal" if enabled else "disabled"
        self.confirm_btn.configure(state=state)
        self.lamp_fail_btn.configure(state=state)

    def _worker(self, sn: str, qr_data: Dict[str, str]) -> None:
        self.result_queue.put(run_setting(input_sn=sn, host=self.host, qr_data=qr_data))

    def _write_final_record(self, record: Dict[str, str]) -> None:
        append_record(self.excel_file, record)
        self.total_count += 1
        self.last_sn = record["input_sn"]
        if record["final_result"] == "PASS":
            self.pass_count += 1
        else:
            self.fail_count += 1
        self._update_counters()

    def _clear_scan_inputs(self) -> None:
        self.qr_var.set("")
        self.sn_var.set("")
        self.current_qr_data = {"raw": "", "sn": "", "mac": "", "wpa": ""}
        self.qr_info_var.set("QR_SN: - | QR_MAC: - | QR_WPA: -")

    def _poll_queue(self) -> None:
        try:
            result = self.result_queue.get_nowait()
        except queue.Empty:
            self.root.after(120, self._poll_queue)
            return
        mode = result["mode"]
        record = result["record"]
        self.last_sn = record["input_sn"]
        if mode == "final":
            self._write_final_record(record)
            self.await_lamp_confirm = False
            self.pending_record = None
            self._set_wait_buttons(False)
            if record["final_result"] == "PASS":
                self._set_status(f"可換下一台 | SN={record['input_sn']}", "#c6efce")
                self._append_ui_log(f"PASS | SN={record['input_sn']} | {record['matched']}")
                self._play_sound(True)
            else:
                self._set_status(f"測試 FAIL | SN={record['input_sn']}", "#ffc7ce")
                self._append_ui_log(f"FAIL | SN={record['input_sn']} | {record['matched']}")
                self._append_ui_log(f"備註: {record['note']}")
                self._play_sound(False)
            self._clear_scan_inputs()
            self._set_busy(False)
            self._refresh_focus()
        else:
            self.pending_record = record
            self.await_lamp_confirm = True
            self._set_wait_buttons(True)
            self._set_status(f"等待燈號確認 | SN={record['input_sn']} | 請確認 GPON 恆亮綠燈", "#9dc3e6")
            self._append_ui_log(f"SN 比對 PASS | Device_SN={record['device_sn']}")
            self._append_ui_log(f"Config PASS | SN={record['input_sn']} | {record['matched']}")
            self._append_ui_log("請依現場燈號選擇：全部 PASS / 換下一台，或 全部 N/A 或燈號 FAIL")
            self._set_busy(False)
            self._play_sound(True)
        self.root.after(120, self._poll_queue)

    def _on_qr_enter(self, event=None) -> None:
        raw = self.qr_var.get().strip()
        parsed = parse_qr_payload(raw)
        self.current_qr_data = {"raw": raw, "sn": parsed["sn"], "mac": parsed["mac"], "wpa": parsed["wpa"]}
        if parsed["sn"]:
            self.sn_var.set(parsed["sn"])
        self.qr_info_var.set(f"QR_SN: {parsed['sn'] or '-'} | QR_MAC: {parsed['mac'] or '-'} | QR_WPA: {parsed['wpa'] or '-'}")
        if parsed["sn"] or parsed["mac"] or parsed["wpa"]:
            self._set_status("QR 解析完成，請確認後開始測試", "#ddebf7")
            self._append_ui_log(f"QR解析 | SN={parsed['sn'] or '-'} | MAC={parsed['mac'] or '-'} | WPA={parsed['wpa'] or '-'}")
            self.root.after(80, lambda: self.start_btn.focus_set())
        else:
            self._set_status("QR 解析失敗，請重新掃碼或手動輸入 SN", "#ffc7ce")
            self._append_ui_log("QR解析失敗：未找到 SN/MAC/WPA")
            self._play_sound(False)
            self.root.after(80, lambda: self.qr_entry.focus_set())

    def _start_test(self) -> None:
        if self.busy or self.await_lamp_confirm:
            return
        sn = self.sn_var.get().strip()
        if not validate_sn(sn):
            self._set_status("SN 格式不正確", "#ffc7ce")
            self._append_ui_log(f"SN 格式錯誤: {sn}")
            self._play_sound(False)
            self._clear_scan_inputs()
            self._refresh_focus()
            return
        if sn_exists(self.excel_file, sn):
            go_on = messagebox.askyesno("SN 重複警告", f"Input_SN {sn} 已存在，是否仍要繼續執行？")
            self._append_ui_log(f"SN 重複檢查: {sn} -> {'繼續' if go_on else '取消'}")
            if not go_on:
                self._set_status(f"取消重複 SN: {sn}", "#fff2cc")
                self._clear_scan_inputs()
                self._refresh_focus()
                return
        self._set_status(f"測試中 | SN={sn}", "#fff2cc")
        self._append_ui_log(f"開始執行 SN: {sn}")
        self.pending_record = None
        self._set_busy(True)
        self._set_wait_buttons(False)
        threading.Thread(target=self._worker, args=(sn, dict(self.current_qr_data)), daemon=True).start()

    def _confirm_next_unit(self) -> None:
        if not self.await_lamp_confirm or not self.pending_record:
            return
        record = dict(self.pending_record)
        record["lamp_check_result"] = "PASS"
        record["final_result"] = "PASS"
        record["note"] = "現場確認 GPON 燈已恆亮綠燈"
        self._write_final_record(record)
        self.await_lamp_confirm = False
        self.pending_record = None
        self._set_wait_buttons(False)
        self._set_status(f"可換下一台 | SN={record['input_sn']}", "#c6efce")
        self._append_ui_log(f"已確認完成，可換下一台 | SN={record['input_sn']}")
        self._clear_scan_inputs()
        self._set_busy(False)
        self._refresh_focus()

    def _record_lamp_fail(self) -> None:
        if not self.await_lamp_confirm or not self.pending_record:
            return
        record = dict(self.pending_record)
        record["lamp_check_result"] = "FAIL"
        record["final_result"] = "FAIL"
        record["note"] = "現場判定 PON/GPON 燈未恆亮綠燈"
        self._write_final_record(record)
        self.await_lamp_confirm = False
        self.pending_record = None
        self._set_wait_buttons(False)
        self._set_status(f"燈號失敗 / 最終 FAIL | SN={record['input_sn']}", "#ffc7ce")
        self._append_ui_log(f"燈號 FAIL | SN={record['input_sn']} | 已記錄最終 FAIL")
        self._clear_scan_inputs()
        self._set_busy(False)
        self._play_sound(False)
        self._refresh_focus()

    def _on_enter(self, event=None) -> None:
        self._start_test()

    def _exit_tool(self) -> None:
        if self.busy and not messagebox.askyesno("確認", "目前測試中，是否仍要離開工具？"):
            return
        if self.await_lamp_confirm and not messagebox.askyesno("確認", "目前尚未按燈號確認，是否仍要離開工具？"):
            return
        self.root.destroy()

def main() -> None:
    logger.info("Application started")
    root = tk.Tk()
    ProductionApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
